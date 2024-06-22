import os
import mysql.connector
import openpyxl
import logging
import time
import requests

logging.basicConfig(
    filename=".Database_writer.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)


def make_dir(folder_name):

    # Combine the current directory with the folder name
    new_path = os.path.join(app_dir, folder_name)
    return new_path


def insert_golfclubs_table(cursor, cnx):

    logging.info("Exporting Golf Clubs Information to GolfClubs Table of MySQL databse")
    golfclubs_excelfile = os.path.join(app_dir, "GolfClubs1.xlsx")
    print(golfclubs_excelfile)
    # Open the Excel file
    wb = openpyxl.load_workbook(golfclubs_excelfile)
    sheet = wb.active
    print(sheet.max_row)
    # Iterate over each row in the sheet (starting from row 2 to skip the header)
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        row = list(row)
        club_name = row[0]
        course_name = row[1]
        address = row[2]
        country = row[5]
        state = row[4]
        suburb = row[3]
        postcode = row[6]
        latitude = row[7]
        longitude = row[8]
        phone = row[10]
        email = row[9]
        golf_cart = row[11]
        club_hire = row[12]
        platform = row[13]
        booking_link = row[14]
        holes_9 = row[15]
        holes_18 = row[16]
        print(club_name)
        print(list(row))
        # Filter the club_name column in the golfclubs table
        query = "SELECT id FROM GolfClubs WHERE club_name = %s"
        cursor.execute(query, (club_name,))
        result = cursor.fetchone()

        # If the club name exists, retrieve the ID
        if result:
            print("ok")
            continue
        else:
            print("no")
            # Insert the data into the GolfClubs table
            query = "INSERT INTO GolfClubs (club_name, course_name, address, country, state, suburb, postcode, latitude, longitude, phone, email, golf_cart, club_hire,  platform, booking_link, 9_holes, 18_holes) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            cursor.execute(
                query,
                (
                    club_name,
                    course_name,
                    address,
                    country,
                    state,
                    suburb,
                    postcode,
                    latitude,
                    longitude,
                    phone,
                    email,
                    golf_cart,
                    club_hire,
                    platform,
                    booking_link,
                    holes_9,
                    holes_18,
                ),
            )

    # Commit the changes to the database
    cnx.commit()


def get_club_id(cursor, cnx, club_name):
    logging.info(
        "Getting the ID Information of  Golf Club that must export to Holes and Booking table : %s",
        club_name,
    )
    # Filter the club_name column in the golfclubs table
    query = "SELECT id FROM GolfClubs WHERE club_name = %s"
    cursor.execute(query, (club_name,))
    result = cursor.fetchone()

    # If the club name exists, retrieve the ID
    if result:
        club_id = result[0]
    else:
        # If the club name does not exist, add it to the club_name column
        query = "INSERT INTO GolfClubs (club_name) VALUES (%s)"
        cursor.execute(query, (club_name,))
        cnx.commit()
        query = "SELECT id FROM GolfClubs WHERE club_name = %s"
        cursor.execute(query, (club_name,))
        result = cursor.fetchone()
        club_id = result[0]
    logging.info("The ID is %s of  %s Golf Club  ", club_id, club_name)
    print(f"club_ID: {club_id}")
    # List all files in the folder
    return club_id


def get_excel_files(folder_name):
    logging.info(
        "Getting the Excel files that must import for Holes and Booking table "
    )
    files = [
        f
        for f in os.listdir(folder_name)
        if os.path.isfile(os.path.join(folder_name, f))
    ]

    # Filter Excel files
    excel_files = [
        f
        for f in files
        if not f.startswith("~") and (f.endswith(".xlsx") or f.endswith(".xls"))
    ]
    logging.info("The Excel files is %s ", excel_files)
    return excel_files


def insert_chrono_data(folder_name, club_name, cursor, cnx):
    logging.info(
        "Writing the data to the Database for %s Golf Club of Chrono Platform from %s Folder",
        club_name,
        folder_name,
    )
    club_id = get_club_id(cursor, cnx, club_name)
    excel_files = get_excel_files(folder_name)
    file = excel_files[0]
    print(excel_files)
    file = os.path.join(chrono_folder, file)
    # Open the Excel file
    wb = openpyxl.load_workbook(file)
    print(file)
    # Iterate over each sheet in the Excel file
    print(len(wb.sheetnames))
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(sheet_name)
        sheet_name = sheet_name.strip()
        # Get the column names

        # Iterate over each data row in the sheet
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            hole_column = row[1]
            print(f"row-num======={row_num}")
            if not hole_column:
                break
            hole_data = hole_column.strip()
            if "9" in hole_data:
                hole_type = 9
            else:
                hole_type = 18
            print(f"hole_type:{hole_type}")

            hole_name = hole_data.strip()
            query = "SELECT id FROM Holes WHERE golf_club_id = %s and hole_name = %s and hole_type = %s"
            cursor.execute(query, (club_id, hole_name, hole_type))
            result = cursor.fetchone()
            # If the club name exists, retrieve the ID
            if result:
                hole_id = result[0]
            else:
                # If the club name does not exist, add it to the club_name column
                query = "INSERT INTO Holes (golf_club_id, hole_type, hole_name ) VALUES (%s, %s, %s)"
                cursor.execute(query, (club_id, hole_type, hole_name))
                cnx.commit()
                query = "SELECT id FROM Holes WHERE golf_club_id = %s and hole_name = %s and hole_type = %s"
                cursor.execute(query, (club_id, hole_name, hole_type))
                result = cursor.fetchone()
                hole_id = result[0]

            time = row[0]
            golfers = int(row[3])
            print(row[2])
            price = row[2]
            try:
                price = float(price)
            except:
                price = float(row[2].replace("$", ""))

            print(price)
            print(f"Hole_ID: {hole_id}")
            logging.info(
                "The hole name of %s Golf Club is %s, Type is %s, ID is %s ",
                club_name,
                hole_name,
                hole_type,
                hole_id,
            )
            query = (
                "SELECT id FROM Bookings WHERE hole_id = %s and date = %s and time = %s"
            )
            cursor.execute(query, (hole_id, sheet_name, time))
            result = cursor.fetchone()
            booking_id = 0
            if result:
                booking_id = result[0]
                # Update the golfers cell value
                query = "UPDATE Bookings SET golfers = %s, price = %s WHERE id = %s"
                cursor.execute(query, (golfers, price, booking_id))
                cnx.commit()
            else:
                # If the club name does not exist, add it to the club_name column
                query = "INSERT INTO Bookings (hole_id, date, time, golfers, price ) VALUES (%s, %s, %s, %s, %s)"
                cursor.execute(query, (hole_id, sheet_name, time, golfers, price))
                cnx.commit()
                query = "SELECT id FROM Bookings WHERE hole_id = %s and date = %s and time = %s"
                cursor.execute(query, (hole_id, sheet_name, time))
                result = cursor.fetchone()
                booking_id = result[0]
            logging.info(
                "The hole name of %s Golf Club is %s, Type is %s, Date is  %s, Time is %s, Available Golfers is %s, Price is %s ",
                club_name,
                hole_name,
                hole_type,
                sheet_name,
                time,
                golfers,
                price,
            )
            print(f"Hole_ID: {booking_id}")


def insert_quick18_data(folder_name, club_name, cursor, cnx):
    logging.info(
        "Writing the data to the Database for %s Golf Club of Quick18 Platfrom from %s Folder",
        club_name,
        folder_name,
    )
    club_id = get_club_id(cursor, cnx, club_name)
    excel_files = get_excel_files(folder_name)
    file = excel_files[0]
    print(excel_files)
    file = os.path.join(quick18_folder, file)
    # Open the Excel file
    wb = openpyxl.load_workbook(file)
    print(file)
    # Iterate over each sheet in the Excel file
    print(len(wb.sheetnames))
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(sheet_name)
        sheet_name = sheet_name.strip()
        # Get the column names
        column_names = [cell.value for cell in sheet[1]]
        print(column_names)
        # Find the index of the column with "adult" in the name
        price = 0

        # Iterate over each data row in the sheet
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            hole_column = row[1]
            hole_data = hole_column.strip()
            if "9" in hole_data:
                hole_type = 9
            else:
                hole_type = 18
            print(f"hole_type:{hole_type}")

            hole_name = hole_data.strip()
            query = "SELECT id FROM Holes WHERE golf_club_id = %s and hole_name = %s and hole_type = %s"
            cursor.execute(query, (club_id, hole_name, hole_type))
            result = cursor.fetchone()
            # If the club name exists, retrieve the ID
            if result:
                hole_id = result[0]
            else:
                # If the club name does not exist, add it to the club_name column
                query = "INSERT INTO Holes (golf_club_id, hole_type, hole_name ) VALUES (%s, %s, %s)"
                cursor.execute(query, (club_id, hole_type, hole_name))
                cnx.commit()
                query = "SELECT id FROM Holes WHERE golf_club_id = %s and hole_name = %s and hole_type = %s"
                cursor.execute(query, (club_id, hole_name, hole_type))
                result = cursor.fetchone()
                hole_id = result[0]

            time = row[0].strip()
            golfers = int(row[3])
            price = row[2].strip().replace("$", "")
            price = float(price)
            print(price)
            print(f"Hole_ID: {hole_id}")
            logging.info(
                "The hole name of %s Golf Club is %s, Type is %s, ID is %s ",
                club_name,
                hole_name,
                hole_type,
                hole_id,
            )
            query = (
                "SELECT id FROM Bookings WHERE hole_id = %s and date = %s and time = %s"
            )
            cursor.execute(query, (hole_id, sheet_name, time))
            result = cursor.fetchone()
            booking_id = 0
            if result:
                booking_id = result[0]
                # Update the golfers cell value
                query = "UPDATE Bookings SET golfers = %s, price = %s WHERE id = %s"
                cursor.execute(query, (golfers, price, booking_id))
                cnx.commit()
            else:
                # If the club name does not exist, add it to the club_name column
                query = "INSERT INTO Bookings (hole_id, date, time, golfers, price ) VALUES (%s, %s, %s, %s, %s)"
                cursor.execute(query, (hole_id, sheet_name, time, golfers, price))
                cnx.commit()
                query = "SELECT id FROM Bookings WHERE hole_id = %s and date = %s and time = %s"
                cursor.execute(query, (hole_id, sheet_name, time))
                result = cursor.fetchone()
                booking_id = result[0]
            logging.info(
                "The hole name of %s Golf Club is %s, Type is %s, Date is  %s, Time is %s, Available Golfers is %s, Price is %s ",
                club_name,
                hole_name,
                hole_type,
                sheet_name,
                time,
                golfers,
                price,
            )
            print(f"Hole_ID: {booking_id}")


def insert_golfbooking_data(folder_name, club_name, cursor, cnx):
    logging.info(
        "Writing the data to the Database for %s Golf Club of Golfbooking Platfrom from %s Folder",
        club_name,
        folder_name,
    )
    club_id = get_club_id(cursor, cnx, club_name)
    excel_files = get_excel_files(folder_name)
    file = excel_files[0]
    print(excel_files)
    file = os.path.join(golfbooking_folder, file)
    # Open the Excel file
    wb = openpyxl.load_workbook(file)
    print(file)
    # Iterate over each sheet in the Excel file
    print(len(wb.sheetnames))
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(sheet_name)
        sheet_name = sheet_name.strip()
        # Get the column names
        column_names = [cell.value for cell in sheet[1]]
        print(column_names)
        # Find the index of the column with "adult" in the name
        price = 0

        # Iterate over each data row in the sheet
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            hole_column = row[1]
            hole_data = hole_column.split("-")
            hole_type = hole_data[1].strip()
            hole_type = int(hole_type.split(" ")[0])
            print(hole_type)
            if hole_type == 9:
                for i, name in enumerate(column_names):
                    if "adult" in str(name).lower() and "9" in str(name).lower():
                        adult_column_index = i
                        break
            else:
                for i, name in enumerate(column_names):
                    if "adult" in str(name).lower() and "18" in str(name).lower():
                        adult_column_index = i
                        break

            print(adult_column_index)

            hole_name = hole_data[0].strip()
            query = "SELECT id FROM Holes WHERE golf_club_id = %s and hole_name = %s and hole_type = %s"
            cursor.execute(query, (club_id, hole_name, hole_type))
            result = cursor.fetchone()
            # If the club name exists, retrieve the ID
            if result:
                hole_id = result[0]
            else:
                # If the club name does not exist, add it to the club_name column
                query = "INSERT INTO Holes (golf_club_id, hole_type, hole_name ) VALUES (%s, %s, %s)"
                cursor.execute(query, (club_id, hole_type, hole_name))
                cnx.commit()
                query = "SELECT id FROM Holes WHERE golf_club_id = %s and hole_name = %s and hole_type = %s"
                cursor.execute(query, (club_id, hole_name, hole_type))
                result = cursor.fetchone()
                hole_id = result[0]

            time = row[0].strip()
            golfers = int(row[3])

            if hole_type == 9:
                price = row[adult_column_index]
            else:
                price = row[adult_column_index]
            price = price.replace("$", "").strip()
            price = float(price)
            print(price)
            print(f"Hole_ID: {hole_id}")
            logging.info(
                "The hole name of %s Golf Club is %s, Type is %s, ID is %s ",
                club_name,
                hole_name,
                hole_type,
                hole_id,
            )
            query = (
                "SELECT id FROM Bookings WHERE hole_id = %s and date = %s and time = %s"
            )
            cursor.execute(query, (hole_id, sheet_name, time))
            result = cursor.fetchone()
            booking_id = 0
            if result:
                booking_id = result[0]
                # Update the golfers cell value
                query = "UPDATE Bookings SET golfers = %s, price = %s WHERE id = %s"
                cursor.execute(query, (golfers, price, booking_id))
                cnx.commit()
            else:
                # If the club name does not exist, add it to the club_name column
                query = "INSERT INTO Bookings (hole_id, date, time, golfers, price ) VALUES (%s, %s, %s, %s, %s)"
                cursor.execute(query, (hole_id, sheet_name, time, golfers, price))
                cnx.commit()
                query = "SELECT id FROM Bookings WHERE hole_id = %s and date = %s and time = %s"
                cursor.execute(query, (hole_id, sheet_name, time))
                result = cursor.fetchone()
                booking_id = result[0]
            logging.info(
                "The hole name of %s Golf Club is %s, Type is %s, Date is  %s, Time is %s, Available Golfers is %s, Price is %s ",
                club_name,
                hole_name,
                hole_type,
                sheet_name,
                time,
                golfers,
                price,
            )
            print(f"Hole_ID: {booking_id}")


def insert_miclub_data(folder_name, club_name, cursor, cnx):
    logging.info(
        "Writing the data to the Database for %s Golf Club of Miclub Platform from %s Folder",
        club_name,
        folder_name,
    )

    club_id = get_club_id(cursor, cnx, club_name)

    excel_files = get_excel_files(folder_name)

    print(f"Excel files in the folder:{excel_files}")
    for file in excel_files:

        hole_data = file.split(".")
        if "9" in hole_data[0]:
            hole_type = 9
        else:
            hole_type = 18

        file = os.path.join(miclub_folder, file)
        # Open the Excel file
        wb = openpyxl.load_workbook(file)
        print(file)
        # Iterate over each sheet in the Excel file
        print(len(wb.sheetnames))
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(sheet_name)
            sheet_name = sheet_name.strip()

            # Iterate over each data row in the sheet
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                print(row[3])
                hole_name = hole_data[0].strip()
                hole_name = f"{hole_name}-{row[1].strip()}"
                query = "SELECT id FROM Holes WHERE golf_club_id = %s and hole_name = %s and hole_type = %s"
                cursor.execute(query, (club_id, hole_name, hole_type))
                result = cursor.fetchone()
                # If the club name exists, retrieve the ID
                if result:
                    hole_id = result[0]
                else:
                    # If the club name does not exist, add it to the club_name column
                    query = "INSERT INTO Holes (golf_club_id, hole_type, hole_name ) VALUES (%s, %s, %s)"
                    cursor.execute(query, (club_id, hole_type, hole_name))
                    cnx.commit()
                    query = "SELECT id FROM Holes WHERE golf_club_id = %s and hole_name = %s and hole_type = %s"
                    cursor.execute(query, (club_id, hole_name, hole_type))
                    result = cursor.fetchone()
                    hole_id = result[0]
                time = row[0].strip()
                golfers = int(row[3])
                price = max(row[4:])
                price = price.replace("$", "").strip()
                price = float(price)
                print(f"Hole_ID: {hole_id}")
                logging.info(
                    "The hole name of %s Golf Club is %s, Type is %s, ID is %s ",
                    club_name,
                    hole_name,
                    hole_type,
                    hole_id,
                )
                query = "SELECT id FROM Bookings WHERE hole_id = %s and date = %s and time = %s"
                cursor.execute(query, (hole_id, sheet_name, time))
                result = cursor.fetchone()
                booking_id = 0
                if result:
                    booking_id = result[0]
                    # Update the golfers cell value
                    query = "UPDATE Bookings SET golfers = %s, price = %s WHERE id = %s"
                    cursor.execute(query, (golfers, price, booking_id))
                    cnx.commit()
                else:
                    # If the club name does not exist, add it to the club_name column
                    query = "INSERT INTO Bookings (hole_id, date, time, golfers, price ) VALUES (%s, %s, %s, %s, %s)"
                    cursor.execute(query, (hole_id, sheet_name, time, golfers, price))
                    cnx.commit()

                    query = "SELECT id FROM Bookings WHERE hole_id = %s and date = %s and time = %s"
                    cursor.execute(query, (hole_id, sheet_name, time))
                    result = cursor.fetchone()
                    booking_id = result[0]
                logging.info(
                    "The hole name of %s Golf Club is %s, Type is %s, Date is  %s, Time is %s, Available Golfers is %s, Price is %s ",
                    club_name,
                    hole_name,
                    hole_type,
                    sheet_name,
                    time,
                    golfers,
                    price,
                )
                print(f"Hole_ID: {booking_id}")

        # Commit the changes to the MySQL database
        cnx.commit()

    """
    # List all directories in the new path
    directories = [d for d in os.listdir(folder_name) if os.path.isdir(os.path.join(folder_name, d))]

    print("Directories in the new path:")
    for directory in directories:
        print(directory)
    """


def delete_database(cursor, cnx):

    # Delete rows from the bookings table where the date column's value is less than the current date
    query = "DELETE FROM Bookings WHERE date < CURDATE()"
    cursor.execute(query)

    # Commit the changes to the database
    cnx.commit()


app_dir = os.path.dirname(__file__)
miclub_folder = make_dir("morningtonGolfclub_Data")
quick18_folder = make_dir("sandringham-Golfclub-data")
golfbooking_folder = make_dir("brighton-Golfclub-data")
chrono_folder = make_dir("chron-Alberpark-Glfclub-data")


if __name__ == "__main__":
    username = "root"
    # password = "y7ZXqfiALL"
    host = "localhost"
    database = "sql12713040"

    # Connect to the MySQL database
    cnx = mysql.connector.connect(
        user=username,
        host=host,
        database=database,
        # user=username, password=password, host=host, database=database
    )

    # Create a cursor object
    cursor = cnx.cursor()

    # Call the insert_golfclubs_table function
    # insert_golfclubs_table(cursor, cnx)
    insert_miclub_data(miclub_folder, "Mornington Golf Club", cursor, cnx)
    insert_golfbooking_data(golfbooking_folder, "Brighton Golf Course", cursor, cnx)
    insert_quick18_data(quick18_folder, "Sandy Golf Links", cursor, cnx)
    insert_chrono_data(chrono_folder, "Albert Park Golf Course", cursor, cnx)
    # delete_database(cursor, cnx)

    # Close the cursor and MySQL connection
    cursor.close()
    cnx.close()
